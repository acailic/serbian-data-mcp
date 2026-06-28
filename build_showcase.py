"""Build comprehensive showcase of Serbian open data potential.

Explores all high and medium potential data categories from data.gov.rs,
fetches real data, creates visualizations, and generates a summary index page.

Categories:
  HIGH: Census, Employment, Air Quality, Municipal Budgets, Real Estate
  MEDIUM: Address Registry, Company Registry, Cross-dataset Analysis
"""

from __future__ import annotations

import asyncio
import json
import logging
import time
from pathlib import Path
from typing import Any

import pandas as pd
import contextlib

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

EXPORT_DIR = Path("exports/showcase")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

results: dict[str, dict[str, Any]] = {}


def save_progress(category: str, info: dict[str, Any]) -> None:
    """Save progress after each category."""
    results[category] = info
    progress_file = EXPORT_DIR / "progress.json"
    progress_file.write_text(json.dumps(results, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    logger.info(f"✅ Saved progress: {category}")


async def fetch_json_url(client, url: str) -> list[dict[str, Any]]:
    """Fetch and parse JSON from a URL with UTF-8 BOM handling."""
    ext = await client._get_external_client()
    resp = await ext.get(url)
    resp.raise_for_status()
    return json.loads(resp.content.decode("utf-8-sig"))


async def category_census(client: Any) -> dict[str, Any]:
    """HIGH 1: Census data — households by municipality (2011 vs 2022)."""
    logger.info("=" * 60)
    logger.info("HIGH 1: Census — Households by municipality")
    logger.info("=" * 60)

    # Fetch 2022 census data
    ds22 = await client.get_dataset("65fc2fc669c6e53ef2b38c50")
    data22 = await fetch_json_url(client, ds22.resources[0].url)

    # Fetch 2011 census data
    ds11 = await client.get_dataset("607fd62a7de272771a0d3842")
    data11 = await fetch_json_url(client, ds11.resources[0].url)

    # Extract municipality-level totals (urban + rural = total, IDBrClDom=0, IDTipNaselja=0)
    def extract_municipalities(data: list[dict]) -> dict[str, float]:
        out: dict[str, float] = {}
        for r in data:
            if r["IDBrClDom"] == 0 and r["IDTipNaselja"] == "0" and r["IDTer"] != "RS":
                out[r["nTer"]] = r["vrednost"]
        return out

    munis22 = extract_municipalities(data22)
    munis11 = extract_municipalities(data11)

    # Build comparison table for matching municipalities
    comparison_rows = []
    for name in sorted(set(munis22.keys()) & set(munis11.keys())):
        v22 = munis22[name]
        v11 = munis11[name]
        change = ((v22 / v11) - 1) * 100 if v11 else 0
        comparison_rows.append(
            {"municipality": name, "households_2011": v11, "households_2022": v22, "change_pct": round(change, 1)}
        )

    df = pd.DataFrame(comparison_rows)
    df = df.sort_values("change_pct", ascending=True)

    # Save data
    csv_path = EXPORT_DIR / "census_households_comparison.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    # Serbia-level totals
    serbia22 = next(
        (
            r["vrednost"]
            for r in data22
            if r["nTer"] == "РЕПУБЛИКА СРБИЈА" and r["IDBrClDom"] == 0 and r["IDTipNaselja"] == "0"
        ),
        0,
    )
    serbia11 = next(
        (
            r["vrednost"]
            for r in data11
            if r["nTer"] == "РЕПУБЛИКА СРБИЈА" and r["IDBrClDom"] == 0 and r["IDTipNaselja"] == "0"
        ),
        0,
    )

    # Household structure breakdown (Serbia 2022)
    structure_rows = []
    for r in data22:
        if r["nTer"] == "РЕПУБЛИКА СРБИЈА" and r["IDTipNaselja"] == "0" and r["IDBrClDom"] != 0:
            structure_rows.append({"category": r["nBrClDom"], "count": r["vrednost"]})
    df_structure = pd.DataFrame(structure_rows)

    # Top 10 biggest losers
    top_decrease = df.head(10)
    # Top 10 biggest gainers
    top_increase = df.tail(10).iloc[::-1]

    info = {
        "status": "success",
        "total_municipalities_2022": len(munis22),
        "total_municipalities_2011": len(munis11),
        "matching_municipalities": len(comparison_rows),
        "serbia_households_2011": serbia11,
        "serbia_households_2022": serbia22,
        "serbia_change_pct": round(((serbia22 / serbia11) - 1) * 100, 1) if serbia11 else None,
        "top_decrease": top_decrease[["municipality", "households_2011", "households_2022", "change_pct"]].to_dict(
            "records"
        ),
        "top_increase": top_increase[["municipality", "households_2011", "households_2022", "change_pct"]].to_dict(
            "records"
        ),
        "household_structure": df_structure.to_dict("records"),
        "csv_path": str(csv_path),
        "data_rows": len(comparison_rows),
    }

    # Save summary
    summary_path = EXPORT_DIR / "census_summary.json"
    summary_path.write_text(json.dumps(info, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    logger.info(f"  Serbia households: {serbia11:,.0f} (2011) → {serbia22:,.0f} (2022) = {info['serbia_change_pct']}%")
    logger.info(f"  Matching municipalities: {len(comparison_rows)}")
    logger.info(f"  Data saved: {csv_path}")

    return info


async def category_employment(client: Any) -> dict[str, Any]:
    """HIGH 2: Employment trends — registered employment since 2015."""
    logger.info("=" * 60)
    logger.info("HIGH 2: Employment — Registered employment trends")
    logger.info("=" * 60)

    # Annual totals since 2015
    ds_total = await client.get_dataset("607fd7dd7de272771a0d3975")
    data_total = await fetch_json_url(client, ds_total.resources[0].url)

    # Extract yearly totals for Serbia (IDModalitetRegZap=0 = total)
    yearly_totals = []
    for r in data_total:
        if r["IDTer"] == "RS" and r["IDModalitetRegZap"] == "0":
            yearly_totals.append({"year": int(r["god"]), "total": r["vrednost"], "label": r["nModalitetRegZap"]})
    df_yearly = pd.DataFrame(yearly_totals).sort_values("year")

    # Extract by employment type for latest year
    by_type_rows = []
    for r in data_total:
        if r["IDTer"] == "RS" and r["god"] == max(
            r["god"] for r in data_total if r["IDTer"] == "RS" and r["IDModalitetRegZap"] == "0"
        ):
            by_type_rows.append(
                {"type_code": r["IDModalitetRegZap"], "type_name": r["nModalitetRegZap"], "value": r["vrednost"]}
            )

    # Monthly data since 2016
    ds_monthly = await client.get_dataset("607fd7e57de272771a0d397a")
    data_monthly = await fetch_json_url(client, ds_monthly.resources[0].url)

    monthly_rows = []
    for r in data_monthly:
        if r["IDTer"] == "RS" and r["IDModalitetRegZap"] == "0":
            monthly_rows.append({"year": int(r["god"]), "month": int(r["mes"]), "total": r["vrednost"]})
    df_monthly = pd.DataFrame(monthly_rows).sort_values(["year", "month"])
    df_monthly["date"] = pd.to_datetime(df_monthly["year"].astype(str) + "-" + df_monthly["month"].astype(str) + "-01")

    # Employment by gender
    ds_gender = await client.get_dataset("607fd7f57de272771a0d3984")
    await fetch_json_url(client, ds_gender.resources[0].url)

    # Save data
    csv_yearly = EXPORT_DIR / "employment_yearly.csv"
    df_yearly.to_csv(csv_yearly, index=False, encoding="utf-8-sig")

    csv_monthly = EXPORT_DIR / "employment_monthly.csv"
    df_monthly.to_csv(csv_monthly, index=False, encoding="utf-8-sig")

    first_year = int(df_yearly["year"].min())
    last_year = int(df_yearly["year"].max())
    first_val = df_yearly[df_yearly["year"] == first_year]["total"].values[0]
    last_val = df_yearly[df_yearly["year"] == last_year]["total"].values[0]
    growth = ((last_val / first_val) - 1) * 100

    info = {
        "status": "success",
        "years_covered": f"{first_year}-{last_year}",
        "first_year_value": first_val,
        "last_year_value": last_val,
        "total_growth_pct": round(growth, 1),
        "yearly_data": df_yearly.to_dict("records"),
        "monthly_data_points": len(df_monthly),
        "employment_types_latest": by_type_rows,
        "csv_yearly": str(csv_yearly),
        "csv_monthly": str(csv_monthly),
    }

    summary_path = EXPORT_DIR / "employment_summary.json"
    summary_path.write_text(json.dumps(info, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    logger.info(f"  Employment: {first_year}: {first_val:,.0f} → {last_year}: {last_val:,.0f} ({growth:+.1f}%)")
    logger.info(f"  Monthly data points: {len(df_monthly)}")
    return info


async def category_air_quality(client: Any) -> dict[str, Any]:
    """HIGH 3: Air quality — PM10 daily data from 40+ stations (2011-2024)."""
    logger.info("=" * 60)
    logger.info("HIGH 3: Air Quality — PM10 daily data from 40+ stations")
    logger.info("=" * 60)

    ds_air = await client.get_dataset("661909571df0e888307e3fa3")

    # Parse 2024 data (most recent complete year)
    ext = await client._get_external_client()
    import openpyxl
    from io import BytesIO

    resp = await ext.get(ds_air.resources[0].url)  # 2024
    resp.raise_for_status()
    wb = openpyxl.load_workbook(BytesIO(resp.content), read_only=True)
    ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(h).strip() if h else "" for h in raw_rows[0]]
    data_rows = raw_rows[1:]

    # Build tidy dataframe
    station_names = [h for h in headers[1:] if h]  # skip first column (date)
    logger.info(f"  Stations found: {len(station_names)}")

    tidy_rows = []
    for row in data_rows:
        if row[0] is None:
            continue
        date_str = str(row[0])
        for i, station in enumerate(station_names):
            val = row[i + 1] if i + 1 < len(row) else None
            if val is not None and val != "" and val != 0:
                with contextlib.suppress(ValueError, TypeError):
                    tidy_rows.append({"date": date_str, "station": station, "pm10": float(val)})

    df = pd.DataFrame(tidy_rows)
    if not df.empty:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"])

    # Compute annual averages per station
    if not df.empty:
        df["year"] = df["date"].dt.year
        station_yearly = df.groupby(["station", "year"])["pm10"].mean().reset_index()
        station_yearly.columns = ["station", "year", "avg_pm10"]

        # 2024 annual averages
        avg_2024 = station_yearly[station_yearly["year"] == 2024].sort_values("avg_pm10", ascending=False)

        # EU limit is 40 µg/m³ annual mean
        numeric_avg = avg_2024["avg_pm10"].apply(lambda x: float(x) if pd.notna(x) else 0)
        stations_exceeding_eu = int((numeric_avg > 40).sum())
        stations_below_eu = int((numeric_avg <= 40).sum())

        # Worst and best stations
        worst = avg_2024.head(5)
        best = avg_2024.tail(5).iloc[::-1]

        # Monthly trend for Serbia (average across all stations)
        df["month"] = df["date"].dt.month
        monthly_avg = df.groupby("month")["pm10"].mean().reset_index()
        monthly_avg.columns = ["month", "avg_pm10"]

        # Save data
        csv_avg = EXPORT_DIR / "air_quality_station_averages_2024.csv"
        avg_2024.to_csv(csv_avg, index=False, encoding="utf-8-sig")

        info = {
            "status": "success",
            "year_analyzed": 2024,
            "total_stations": len(station_names),
            "data_points": len(df),
            "date_range": f"{df['date'].min()} to {df['date'].max()}",
            "eu_limit_ugm3": 40,
            "stations_exceeding_eu_limit": stations_exceeding_eu,
            "stations_below_eu_limit": stations_below_eu,
            "worst_5_stations": worst[["station", "avg_pm10"]].round(1).to_dict("records"),
            "best_5_stations": best[["station", "avg_pm10"]].round(1).to_dict("records"),
            "monthly_pattern": monthly_avg.round(1).to_dict("records"),
            "station_names": station_names,
            "csv_path": str(csv_avg),
        }
    else:
        info = {"status": "failed", "reason": "No data parsed from 2024 XLSX"}

    summary_path = EXPORT_DIR / "air_quality_summary.json"
    summary_path.write_text(json.dumps(info, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    if info["status"] == "success":
        logger.info(f"  Stations: {info['total_stations']}, Data points: {info['data_points']}")
        logger.info(
            f"  Exceeding EU limit (40µg/m³): {info['stations_exceeding_eu_limit']} of {info['total_stations']}"
        )
        logger.info(
            f"  Worst: {info['worst_5_stations'][0]['station']} ({info['worst_5_stations'][0]['avg_pm10']}µg/m³)"
        )
        logger.info(f"  Best: {info['best_5_stations'][0]['station']} ({info['best_5_stations'][0]['avg_pm10']}µg/m³)")

    return info


async def category_budgets(client: Any) -> dict[str, Any]:
    """HIGH 4: Municipal budgets — analysis of budget data availability."""
    logger.info("=" * 60)
    logger.info("HIGH 4: Municipal Budgets — Data availability analysis")
    logger.info("=" * 60)

    # Scan all budget datasets
    all_budgets = []
    page = 1
    while True:
        result = await client.search_datasets(query="Буџет", page_size=100, page=page)
        if not result.datasets:
            break
        for ds in result.datasets:
            org_name = ds.organization.name if ds.organization else "Unknown"
            formats = sorted({r.format for r in ds.resources if r.format})
            all_budgets.append(
                {
                    "id": ds.id,
                    "title": ds.title,
                    "organization": org_name,
                    "formats": formats,
                    "resource_count": len(ds.resources),
                    "tags": ds.tags or [],
                }
            )
        if not result.has_next:
            break
        page += 1

    # Analyze by municipality
    from collections import Counter

    org_counts = Counter(b["organization"] for b in all_budgets)
    format_counts = Counter(f for b in all_budgets for f in b["formats"])
    year_mentions = []
    for b in all_budgets:
        import re

        years = re.findall(r"(20\d{2})", b["title"])
        year_mentions.extend(years)
    year_counts = Counter(year_mentions)

    # Try to download one budget file to see structure
    sample_data_info = None
    for b in all_budgets[:50]:
        if "csv" in b["formats"]:
            try:
                ds = await client.get_dataset(b["id"])
                if ds.resources:
                    res = ds.resources[0]
                    ext = await client._get_external_client()
                    resp = await ext.get(res.url)
                    resp.raise_for_status()
                    content = resp.text[:2000]
                    sample_data_info = {
                        "title": b["title"],
                        "organization": b["organization"],
                        "format": res.format,
                        "first_2000_chars": content,
                    }
                    break
            except Exception:
                continue

    info = {
        "status": "success",
        "total_budget_datasets": len(all_budgets),
        "unique_municipalities": len(org_counts),
        "top_municipalities_by_count": org_counts.most_common(15),
        "format_distribution": dict(format_counts.most_common()),
        "years_covered": dict(sorted(year_counts.items())),
        "sample_data_preview": sample_data_info,
    }

    summary_path = EXPORT_DIR / "budgets_summary.json"
    summary_path.write_text(json.dumps(info, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    logger.info(f"  Total budget datasets: {len(all_budgets)}")
    logger.info(f"  Unique municipalities: {len(org_counts)}")
    logger.info(f"  Years: {dict(sorted(year_counts.items()))}")
    return info


async def category_real_estate(client: Any) -> dict[str, Any]:
    """HIGH 5: Real estate prices — property values by municipality."""
    logger.info("=" * 60)
    logger.info("HIGH 5: Real Estate — Average property prices")
    logger.info("=" * 60)

    # Find property price datasets
    result = await client.search_datasets("непокретности", page_size=50)

    datasets_info = []
    for ds in result.datasets:
        org_name = ds.organization.name if ds.organization else "Unknown"
        formats = sorted({r.format for r in ds.resources if r.format})
        datasets_info.append(
            {
                "id": ds.id,
                "title": ds.title,
                "organization": org_name,
                "formats": formats,
                "resource_count": len(ds.resources),
            }
        )

    # Try to download the national-level price dataset
    national_price_data = None
    for ds_info in datasets_info:
        if "ПРОСЕЧНИХ" in ds_info["title"] or "Просечн" in ds_info["title"]:
            try:
                ds = await client.get_dataset(ds_info["id"])
                if ds.resources:
                    res = ds.resources[0]
                    ext = await client._get_external_client()
                    resp = await ext.get(res.url)
                    resp.raise_for_status()

                    import openpyxl
                    from io import BytesIO

                    wb = openpyxl.load_workbook(BytesIO(resp.content), read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    wb.close()

                    # Parse structure
                    if len(rows) > 1:
                        headers = [str(h).strip() if h else "" for h in rows[0]]
                        sample = [str(r[:8]) if r else "" for r in rows[1:6]]
                        national_price_data = {
                            "dataset_title": ds_info["title"],
                            "organization": ds_info["organization"],
                            "headers": headers[:10],
                            "sample_rows": sample,
                            "total_rows": len(rows) - 1,
                        }
                        break
            except Exception as e:
                logger.warning(f"  Failed to parse {ds_info['title']}: {e}")
                continue

    info = {
        "status": "success",
        "total_datasets": len(datasets_info),
        "datasets": datasets_info[:15],
        "national_price_preview": national_price_data,
    }

    summary_path = EXPORT_DIR / "real_estate_summary.json"
    summary_path.write_text(json.dumps(info, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    logger.info(f"  Real estate datasets: {len(datasets_info)}")
    if national_price_data:
        logger.info(
            f"  National data: {national_price_data['total_rows']} rows, cols: {national_price_data['headers']}"
        )
    return info


async def category_registries(client: Any) -> dict[str, Any]:
    """MEDIUM 6: Public registries — address, company, energy permits."""
    logger.info("=" * 60)
    logger.info("MEDIUM 6: Public Registries — Address, Company, Energy")
    logger.info("=" * 60)

    registries = {
        "Адресни регистар": "5fcf5ca67de272573b9c2722",
        "АПИ за Регистар привредних друштава": "68000c424d29e8a004f93e04",
        "АПИ за Регистар задужбина": "686372744ad4298af55abeff",
        "Регистар издатих енергетских дозвола": "58a1aeeecbe3c810aa69cd3f",
        "Регистар повлашћених произвођача": "58ac4098cbe3c810aa69cd4b",
    }

    registry_info = {}
    for name, ds_id in registries.items():
        try:
            ds = await client.get_dataset(ds_id)
            org = ds.organization.name if ds.organization else "N/A"
            resources = [{"title": r.title, "format": r.format, "size": r.size} for r in ds.resources]
            registry_info[name] = {"organization": org, "resources": resources, "description": ds.description[:200]}
        except Exception as e:
            registry_info[name] = {"error": str(e)}

    # Try company registry API
    company_api_info = None
    ds_co = await client.get_dataset("68000c424d29e8a004f93e04")
    for r in ds_co.resources:
        if r.format == "json":
            try:
                ext = await client._get_external_client()
                resp = await ext.get(r.url, follow_redirects=True)
                resp.raise_for_status()
                company_api_info = {
                    "url": r.url,
                    "status_code": resp.status_code,
                    "content_type": resp.headers.get("content-type", ""),
                    "content_preview": resp.text[:1000],
                }
            except Exception as e:
                company_api_info = {"url": r.url, "error": str(e)}
            break

    info = {
        "status": "success",
        "registries": registry_info,
        "company_api": company_api_info,
    }

    summary_path = EXPORT_DIR / "registries_summary.json"
    summary_path.write_text(json.dumps(info, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    logger.info(f"  Registries analyzed: {len(registry_info)}")
    if company_api_info:
        logger.info(f"  Company API: {company_api_info.get('status_code', 'error')}")
    return info


async def category_cross_analysis(client: Any) -> dict[str, Any]:
    """MEDIUM 7: Cross-dataset analysis — employment + demographics correlation."""
    logger.info("=" * 60)
    logger.info("MEDIUM 7: Cross-dataset Analysis — Employment & Demographics")
    logger.info("=" * 60)

    # Get employment by municipality
    ds_emp_mun = await client.get_dataset("607fd7da7de272771a0d3973")
    data_emp = await fetch_json_url(client, ds_emp_mun.resources[0].url)

    # Get latest year employment by municipality
    emp_years = sorted({r["god"] for r in data_emp if r["IDTer"] != "RS"}, reverse=True)
    latest_year = emp_years[0] if emp_years else "2023"

    emp_by_muni = {}
    for r in data_emp:
        if r["god"] == latest_year and r["IDTer"] != "RS" and r.get("IDModalitetRegZap", "0") == "0":
            emp_by_muni[r["nTer"]] = r["vrednost"]

    # Get census households by municipality
    ds_census = await client.get_dataset("65fc2fc669c6e53ef2b38c50")
    data_census = await fetch_json_url(client, ds_census.resources[0].url)

    households_by_muni = {}
    for r in data_census:
        if r["IDBrClDom"] == 0 and r["IDTipNaselja"] == "0" and r["IDTer"] != "RS":
            households_by_muni[r["nTer"]] = r["vrednost"]

    # Match and compute employment per household ratio
    cross_rows = []
    for muni in sorted(set(emp_by_muni.keys()) & set(households_by_muni.keys())):
        emp = emp_by_muni[muni]
        hh = households_by_muni[muni]
        ratio = emp / hh if hh else 0
        cross_rows.append(
            {"municipality": muni, "employment": emp, "households": hh, "emp_per_household": round(ratio, 2)}
        )

    df = pd.DataFrame(cross_rows)
    df = df.sort_values("emp_per_household", ascending=False)

    csv_path = EXPORT_DIR / "cross_analysis_employment_households.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    info = {
        "status": "success",
        "employment_year": latest_year,
        "census_year": "2022",
        "matching_municipalities": len(cross_rows),
        "top_10_emp_per_household": df.head(10).to_dict("records"),
        "bottom_10_emp_per_household": df.tail(10).to_dict("records"),
        "serbia_avg_emp_per_household": round(df["emp_per_household"].mean(), 2),
        "csv_path": str(csv_path),
    }

    summary_path = EXPORT_DIR / "cross_analysis_summary.json"
    summary_path.write_text(json.dumps(info, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    logger.info(f"  Matching municipalities: {len(cross_rows)}")
    logger.info(f"  Avg employment per household: {info['serbia_avg_emp_per_household']}")
    logger.info(f"  Highest: {df.iloc[0]['municipality']} ({df.iloc[0]['emp_per_household']})")
    logger.info(f"  Lowest: {df.iloc[-1]['municipality']} ({df.iloc[-1]['emp_per_household']})")
    return info


async def category_municipal_data_availability(client: Any) -> dict[str, Any]:
    """MEDIUM 8: Municipal data availability — which municipalities publish what."""
    logger.info("=" * 60)
    logger.info("MEDIUM 8: Municipal Data Availability Analysis")
    logger.info("=" * 60)

    # Collect all datasets with org info
    all_datasets = []
    page = 1
    while True:
        result = await client.search_datasets(query="", page_size=100, page=page)
        if not result.datasets:
            break
        for ds in result.datasets:
            org = ds.organization.name if ds.organization else "Unknown"
            tags = ds.tags or []
            formats = sorted({r.format for r in ds.resources if r.format})
            all_datasets.append(
                {
                    "id": ds.id,
                    "title": ds.title,
                    "org": org,
                    "tags": tags,
                    "formats": formats,
                    "modified": str(ds.modified_at)[:10] if ds.modified_at else "",
                }
            )
        if not result.has_next or page > 35:
            break
        page += 1

    # Classify organizations
    from collections import Counter

    org_dataset_count = Counter(d["org"] for d in all_datasets)
    org_format_profile: dict[str, Counter] = {}
    for d in all_datasets:
        if d["org"] not in org_format_profile:
            org_format_profile[d["org"]] = Counter()
        for f in d["formats"]:
            org_format_profile[d["org"]][f] += 1

    # Top publishers
    top_pubs = [{"org": org, "datasets": count} for org, count in org_dataset_count.most_common(20)]

    # Tag distribution
    all_tags = Counter()
    for d in all_datasets:
        for t in d["tags"]:
            all_tags[t] += 1

    # Format distribution
    all_formats = Counter()
    for d in all_datasets:
        for f in d["formats"]:
            all_formats[f] += 1

    # Government vs non-government classification
    gov_keywords = [
        "општина",
        "град ",
        "министарство",
        "завод",
        "институт",
        "управа",
        "агенција",
        "републички",
        "покрајински",
    ]
    gov_datasets = sum(1 for d in all_datasets if any(kw in d["org"].lower() for kw in gov_keywords))
    non_gov = len(all_datasets) - gov_datasets

    info = {
        "status": "success",
        "total_datasets_scanned": len(all_datasets),
        "government_datasets": gov_datasets,
        "non_government_datasets": non_gov,
        "top_publishers": top_pubs,
        "format_distribution": dict(all_formats.most_common()),
        "tag_distribution": dict(all_tags.most_common(20)),
        "pages_scanned": page,
    }

    summary_path = EXPORT_DIR / "data_availability_summary.json"
    summary_path.write_text(json.dumps(info, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    logger.info(f"  Datasets scanned: {len(all_datasets)}")
    logger.info(f"  Government: {gov_datasets}, Non-government: {non_gov}")
    logger.info(f"  Top publisher: {top_pubs[0]}")
    return info


async def build_index_html() -> None:
    """Build a comprehensive index HTML page summarizing all findings."""
    logger.info("Building index page...")

    # Load all summaries
    summaries: dict[str, Any] = {}
    for f in EXPORT_DIR.glob("*_summary.json"):
        name = f.stem.replace("_summary", "")
        summaries[name] = json.loads(f.read_text(encoding="utf-8"))

    html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Serbian Open Data — What's Possible</title>
    <style>
:root {
    --bg: #0f1117;
    --bg-card: #1a1d28;
    --text: #e8e8ed;
    --text-dim: #8b8d97;
    --accent: #e63946;
    --accent2: #1d8cf8;
    --accent3: #f4a261;
    --accent4: #2a9d8f;
    --border: rgba(255,255,255,0.08);
}
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
    background: var(--bg);
    color: var(--text);
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
    line-height: 1.7;
}
.container { max-width: 1100px; margin: 0 auto; padding: 40px 24px; }
h1 {
    font-size: 2.4rem;
    font-weight: 800;
    text-align: center;
    margin-bottom: 8px;
    background: linear-gradient(135deg, #e63946, #f4a261);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
}
.subtitle { text-align: center; color: var(--text-dim); margin-bottom: 48px; font-size: 1.1rem; }
.section {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 28px;
    margin-bottom: 24px;
}
.section h2 {
    font-size: 1.4rem;
    margin-bottom: 4px;
    display: flex;
    align-items: center;
    gap: 10px;
}
.badge {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 0.7rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}
.badge-high { background: #e63946; color: white; }
.badge-med { background: #f4a261; color: #0f1117; }
.section .desc { color: var(--text-dim); margin-bottom: 16px; font-size: 0.95rem; }
.big-num {
    font-size: 2.2rem;
    font-weight: 800;
    color: var(--accent);
}
.big-num.blue { color: var(--accent2); }
.big-num.green { color: var(--accent4); }
.big-num.gold { color: var(--accent3); }
.stats {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 16px;
    margin: 16px 0;
}
.stat-box {
    background: rgba(255,255,255,0.03);
    border-radius: 8px;
    padding: 16px;
    text-align: center;
}
.stat-box .label { color: var(--text-dim); font-size: 0.8rem; margin-bottom: 4px; }
.findings { margin-top: 16px; }
.findings h3 { font-size: 1rem; margin-bottom: 8px; color: var(--accent3); }
.findings ul { padding-left: 20px; color: var(--text-dim); }
.findings li { margin-bottom: 6px; }
.findings strong { color: var(--text); }
table {
    width: 100%;
    border-collapse: collapse;
    margin: 12px 0;
    font-size: 0.85rem;
}
th, td {
    padding: 8px 12px;
    text-align: left;
    border-bottom: 1px solid var(--border);
}
th { color: var(--text-dim); font-weight: 600; }
td { color: var(--text); }
.negative { color: #e63946; }
.positive { color: #2a9d8f; }
.tag {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 0.75rem;
    background: rgba(29,140,248,0.15);
    color: var(--accent2);
    margin: 2px;
}
.footer { text-align: center; color: var(--text-dim); margin-top: 48px; font-size: 0.85rem; }
    </style>
</head>
<body>
<div class="container">
    <h1>🇷🇸 Serbian Open Data: What's Possible</h1>
    <p class="subtitle">
        Investigation of data.gov.rs — 3,443 datasets from 184 organizations<br>
        Real data, real analysis, real potential
    </p>
"""

    # Portal overview
    availability = summaries.get("data_availability", {})
    html += f"""
    <div class="section">
        <h2>📊 Portal Overview</h2>
        <p class="desc">National open data portal data.gov.rs — scope and composition</p>
        <div class="stats">
            <div class="stat-box">
                <div class="big-num">{availability.get("total_datasets_scanned", 3443):,}</div>
                <div class="label">Total Datasets</div>
            </div>
            <div class="stat-box">
                <div class="big-num blue">184</div>
                <div class="label">Organizations</div>
            </div>
            <div class="stat-box">
                <div class="big-num green">{availability.get("government_datasets", "?")}</div>
                <div class="label">Government Datasets</div>
            </div>
            <div class="stat-box">
                <div class="big-num gold">{availability.get("non_government_datasets", "?")}</div>
                <div class="label">Non-Government</div>
            </div>
        </div>
        <div class="findings">
            <h3>Key Findings</h3>
            <ul>
                <li>Portal is dominated by <strong>municipal budget uploads</strong> — one municipality (Рашка) has 245 datasets alone</li>
                <li>Many datasets are from <strong>private businesses</strong> uploading price lists (noise data)</li>
                <li>The API search only works with <strong>Cyrillic Serbian</strong> — English queries return 0 results</li>
                <li>The <strong>Statistical Office (РЗС)</strong> has the highest-quality structured data (JSON/CSV, 34K+ rows)</li>
                <li><strong>15+ years of daily air quality data</strong> from 40+ stations is a standout dataset</li>
            </ul>
        </div>
    </div>
"""

    # HIGH 1: Census
    census = summaries.get("census", {})
    if census.get("status") == "success":
        html += f"""
    <div class="section">
        <h2><span class="badge badge-high">HIGH</span> Census — Households by Municipality</h2>
        <p class="desc">Popis 2022 vs Popis 2011: 2.5M+ households across {census.get("matching_municipalities", 0)} municipalities</p>
        <div class="stats">
            <div class="stat-box">
                <div class="big-num">{census.get("serbia_households_2022", 0):,.0f}</div>
                <div class="label">Households 2022</div>
            </div>
            <div class="stat-box">
                <div class="big-num blue">{census.get("serbia_change_pct", "?")}%</div>
                <div class="label">Change Since 2011</div>
            </div>
            <div class="stat-box">
                <div class="big-num gold">{census.get("matching_municipalities", 0)}</div>
                <div class="label">Municipalities Compared</div>
            </div>
        </div>
        <div class="findings">
            <h3>What's Possible</h3>
            <ul>
                <li><strong>Municipality-level comparison</strong> of household growth/decline over 11 years</li>
                <li><strong>Choropleth map</strong> of Serbia showing household change by district (25 okruga)</li>
                <li><strong>Household structure analysis</strong> — 1-person vs multi-person households</li>
                <li><strong>Slope chart</strong> ranking municipalities by population change</li>
                <li>Urban vs rural settlement breakdown available</li>
            </ul>
            <h3>Biggest Decline</h3>
            <table><tr><th>Municipality</th><th>2011</th><th>2022</th><th>Change</th></tr>"""
        for row in census.get("top_decrease", [])[:5]:
            html += f'<tr><td>{row["municipality"]}</td><td>{row["households_2011"]:,.0f}</td><td>{row["households_2022"]:,.0f}</td><td class="negative">{row["change_pct"]}%</td></tr>'
        html += """</table></div>
    </div>"""

    # HIGH 2: Employment
    emp = summaries.get("employment", {})
    if emp.get("status") == "success":
        html += f"""
    <div class="section">
        <h2><span class="badge badge-high">HIGH</span> Employment — Registered Employment Trends</h2>
        <p class="desc">Monthly and annual registered employment since 2015 — {emp.get("monthly_data_points", 0)} data points</p>
        <div class="stats">
            <div class="stat-box">
                <div class="big-num">{emp.get("last_year_value", 0):,.0f}</div>
                <div class="label">Employed ({emp.get("years_covered", "?").split("-")[-1]})</div>
            </div>
            <div class="stat-box">
                <div class="big-num green">+{emp.get("total_growth_pct", "?")}%</div>
                <div class="label">Growth Since {emp.get("years_covered", "?").split("-")[0]}</div>
            </div>
            <div class="stat-box">
                <div class="big-num gold">{emp.get("monthly_data_points", 0)}</div>
                <div class="label">Monthly Data Points</div>
            </div>
        </div>
        <div class="findings">
            <h3>What's Possible</h3>
            <ul>
                <li><strong>Time series line chart</strong> of employment growth 2015-2024 (10-year trend)</li>
                <li><strong>Seasonal pattern analysis</strong> — monthly data reveals hiring cycles</li>
                <li><strong>Employment type breakdown</strong> — legal entities vs entrepreneurs vs farmers</li>
                <li><strong>Forecasting</strong> — project employment to 2030 with linear regression</li>
                <li><strong>Municipality-level</strong> employment data available for cross-analysis</li>
            </ul>
            <h3>Yearly Trend</h3>
            <table><tr><th>Year</th><th>Total Employed</th><th>YoY Change</th></tr>"""
        yearly = emp.get("yearly_data", [])
        for i, row in enumerate(yearly):
            yoy = ""
            if i > 0:
                prev = yearly[i - 1]["total"]
                if prev:
                    pct = ((row["total"] / prev) - 1) * 100
                    cls = "positive" if pct >= 0 else "negative"
                    yoy = f'<span class="{cls}">{pct:+.1f}%</span>'
            html += f"<tr><td>{row['year']}</td><td>{row['total']:,.0f}</td><td>{yoy}</td></tr>"
        html += """</table></div>
    </div>"""

    # HIGH 3: Air Quality
    air = summaries.get("air_quality", {})
    if air.get("status") == "success":
        worst = air.get("worst_5_stations", [])
        best = air.get("best_5_stations", [])
        html += f"""
    <div class="section">
        <h2><span class="badge badge-high">HIGH</span> Air Quality — Daily PM₁₀ from {air.get("total_stations", 0)} Stations</h2>
        <p class="desc">
            Verified daily PM₁₀ data, {air.get("data_points", 0):,} measurements, {air.get("year_analyzed", 2024)} annual analysis<br>
            Years available: 2011-2024 (14 years!) — EU limit: 40 µg/m³ annual mean
        </p>
        <div class="stats">
            <div class="stat-box">
                <div class="big-num">{air.get("total_stations", 0)}</div>
                <div class="label">Monitoring Stations</div>
            </div>
            <div class="stat-box">
                <div class="big-num" style="color: #e63946">{air.get("stations_exceeding_eu_limit", 0)}</div>
                <div class="label">Exceed EU Limit</div>
            </div>
            <div class="stat-box">
                <div class="big-num green">{air.get("stations_below_eu_limit", 0)}</div>
                <div class="label">Below EU Limit</div>
            </div>
        </div>
        <div class="findings">
            <h3>What's Possible</h3>
            <ul>
                <li><strong>14-year time series</strong> of air quality — unmatched temporal depth</li>
                <li><strong>Station-level heatmap</strong> showing daily PM₁₀ across all 40+ stations</li>
                <li><strong>Seasonal analysis</strong> — winter heating vs summer pollution patterns</li>
                <li><strong>City ranking</strong> — which cities breathe the worst air</li>
                <li><strong>Before/after comparison</strong> — policy intervention analysis</li>
                <li><strong>Choropleth map</strong> of Serbia colored by air quality</li>
            </ul>
            <h3>Worst Air Quality (2024 Annual Mean PM₁₀)</h3>
            <table><tr><th>Station</th><th>Avg PM₁₀ (µg/m³)</th><th>vs EU Limit</th></tr>"""
        for row in worst[:5]:
            pct = row["avg_pm10"] / 40 * 100
            html += f'<tr><td>{row["station"]}</td><td>{row["avg_pm10"]}</td><td class="negative">{pct:.0f}% of limit</td></tr>'
        html += """</table>
            <h3>Cleanest Air</h3>
            <table><tr><th>Station</th><th>Avg PM₁₀ (µg/m³)</th></tr>"""
        for row in best[:5]:
            html += f'<tr><td>{row["station"]}</td><td class="positive">{row["avg_pm10"]}</td></tr>'
        html += """</table></div>
    </div>"""

    # HIGH 4: Budgets
    budgets = summaries.get("budgets", {})
    if budgets.get("status") == "success":
        top_muns = budgets.get("top_municipalities_by_count", [])
        html += f"""
    <div class="section">
        <h2><span class="badge badge-high">HIGH</span> Municipal Budgets — {budgets.get("total_budget_datasets", 0)} Budget Datasets</h2>
        <p class="desc">
            Budget execution data from {budgets.get("unique_municipalities", 0)} municipalities.
            Multi-year data in XLSX/CSV format.
        </p>
        <div class="stats">
            <div class="stat-box">
                <div class="big-num">{budgets.get("total_budget_datasets", 0)}</div>
                <div class="label">Budget Datasets</div>
            </div>
            <div class="stat-box">
                <div class="big-num blue">{budgets.get("unique_municipalities", 0)}</div>
                <div class="label">Municipalities</div>
            </div>
            <div class="stat-box">
                <div class="big-num gold">XLSX</div>
                <div class="label">Primary Format</div>
            </div>
        </div>
        <div class="findings">
            <h3>What's Possible</h3>
            <ul>
                <li><strong>Budget comparison</strong> across municipalities — who spends most per capita</li>
                <li><strong>Revenue vs expenditure</strong> analysis — balanced vs deficit municipalities</li>
                <li><strong>Multi-year trends</strong> — how budgets evolved 2019-2026</li>
                <li><strong>Budget structure</strong> — revenue sources, spending categories (using Sankey diagrams)</li>
                <li><strong>Format challenge</strong>: XLSX files have varied structures — needs per-municipality parsing</li>
            </ul>
            <h3>Top Publishers</h3>
            <table><tr><th>Municipality</th><th>Datasets</th></tr>"""
        for org, count in top_muns[:8]:
            html += f"<tr><td>{org}</td><td>{count}</td></tr>"
        html += """</table></div>
    </div>"""

    # HIGH 5: Real Estate
    re_data = summaries.get("real_estate", {})
    if re_data.get("status") == "success":
        npp = re_data.get("national_price_preview", {})
        html += f"""
    <div class="section">
        <h2><span class="badge badge-high">HIGH</span> Real Estate — Property Prices</h2>
        <p class="desc">{re_data.get("total_datasets", 0)} datasets with average square meter prices for property tax assessment</p>
        <div class="findings">
            <h3>What's Available</h3>
            <ul>"""
        for ds in re_data.get("datasets", [])[:10]:
            html += f"<li><strong>{ds['title']}</strong> — {ds['organization']} ({', '.join(ds['formats'])})</li>"
        html += """</ul>
            <h3>What's Possible</h3>
            <ul>
                <li><strong>Property price comparison</strong> across municipalities</li>
                <li><strong>Price trend analysis</strong> — yearly price changes per municipality</li>
                <li><strong>Choropleth map</strong> of Serbia colored by property prices</li>
                <li>Data format is <strong>tax assessment prices</strong> (not market prices) — conservative baseline</li>
            </ul>"""
        if npp:
            html += f"""<h3>Data Structure (National)</h3>
            <p style="color:var(--text-dim);font-size:0.9rem">{npp.get("total_rows", "?")} rows, columns: {", ".join(npp.get("headers", []))}</p>"""
        html += """</div>
    </div>"""

    # MEDIUM 6: Registries
    regs = summaries.get("registries", {})
    if regs.get("status") == "success":
        api = regs.get("company_api", {})
        html += """
    <div class="section">
        <h2><span class="badge badge-med">MEDIUM</span> Public Registries — Address, Company, Energy</h2>
        <p class="desc">Government registries available as APIs and bulk downloads</p>
        <div class="findings">
            <h3>Available Registries</h3>"""
        for name, info in regs.get("registries", {}).items():
            status = "✅" if not info.get("error") else "❌"
            formats = ", ".join(r["format"] for r in info.get("resources", []))
            html += f"<li>{status} <strong>{name}</strong> — {info.get('organization', '?')} ({formats})</li>"
        if api:
            if api.get("status_code"):
                html += f"""
            <h3>Company Registry API</h3>
            <p>Live API endpoint — returns {api.get("content_type", "?")} — status {api.get("status_code")}</p>
            <p style="color:var(--text-dim);font-size:0.85rem">Preview: {api.get("content_preview", "")[:200]}...</p>"""
            else:
                html += f'<p style="color:var(--accent)">Company API error: {api.get("error", "?")}</p>'
        html += """
            <h3>What's Possible</h3>
            <ul>
                <li><strong>Address registry with GeoPKG</strong> — geospatial data of all Serbian addresses</li>
                <li><strong>Company registry API</strong> — real-time access to business registration data</li>
                <li><strong>Energy permit tracking</strong> — who's generating renewable energy</li>
            </ul>
        </div>
    </div>"""

    # MEDIUM 7: Cross-analysis
    cross = summaries.get("cross_analysis", {})
    if cross.get("status") == "success":
        html += f"""
    <div class="section">
        <h2><span class="badge badge-med">MEDIUM</span> Cross-Dataset Analysis — Employment × Households</h2>
        <p class="desc">
            Matching {cross.get("matching_municipalities", 0)} municipalities across census (2022) and employment ({cross.get("employment_year", "?")}) data.
            Employment per household ratio reveals economic activity patterns.
        </p>
        <div class="stats">
            <div class="stat-box">
                <div class="big-num">{cross.get("serbia_avg_emp_per_household", "?")}</div>
                <div class="label">Avg Employment per Household</div>
            </div>
            <div class="stat-box">
                <div class="big-num blue">{cross.get("matching_municipalities", 0)}</div>
                <div class="label">Municipalities Matched</div>
            </div>
        </div>
        <div class="findings">
            <h3>Top Employment-per-Household Municipalities</h3>
            <table><tr><th>Municipality</th><th>Employment</th><th>Households</th><th>Emp/HH</th></tr>"""
        for row in cross.get("top_10_emp_per_household", [])[:5]:
            html += f"<tr><td>{row['municipality']}</td><td>{row['employment']:,.0f}</td><td>{row['households']:,.0f}</td><td>{row['emp_per_household']}</td></tr>"
        html += """
            </table>
            <h3>What's Possible</h3>
            <ul>
                <li><strong>Economic activity mapping</strong> — which municipalities are employment hubs</li>
                <li><strong>Choropleth map</strong> of employment density per household</li>
                <li><strong>Radar chart</strong> comparing top municipalities across multiple indicators</li>
                <li><strong>Correlation analysis</strong> — does household size predict employment?</li>
            </ul>
        </div>
    </div>"""

    # Summary / recommendations
    html += """
    <div class="section">
        <h2>🎯 Summary & Recommendations</h2>
        <div class="findings">
            <h3>Ready for Production (Start Here)</h3>
            <ul>
                <li><strong>Air Quality Dashboard</strong> — 14 years, 40+ stations, daily data. Highest ROI dataset on the portal. Build animated time-series with city toggle.</li>
                <li><strong>Employment Trends Dashboard</strong> — 10 years of monthly data. Line charts with forecasting to 2030. Seasonal decomposition.</li>
                <li><strong>Census Comparison Map</strong> — Choropleth of household change 2011→2022. Slope chart for ranking changes. Infographic with demographic insights.</li>
            </ul>
            <h3>Worth Exploring (Moderate Effort)</h3>
            <ul>
                <li><strong>Municipal Budget Analysis</strong> — Requires per-municipality XLSX parsing (varied structures). Sankey diagrams of budget flows. Comparison across municipalities.</li>
                <li><strong>Real Estate Price Map</strong> — Tax assessment prices as baseline. Choropleth by municipality. Year-over-year trend analysis.</li>
                <li><strong>Cross-Dataset Insights</strong> — Employment × demographics, air quality × budgets, registries × economic data.</li>
            </ul>
            <h3>Technical Improvements Needed</h3>
            <ul>
                <li><strong>Fix resource ID lookup</strong> — The `_find_resource` method only scans 150 recent datasets. Must search by parent dataset ID first.</li>
                <li><strong>UTF-8 BOM handling</strong> — RZS JSON endpoint returns BOM, parser must use `utf-8-sig`.</li>
                <li><strong>Search language</strong> — Portal API only supports Cyrillic. The bilingual intelligent search is essential for usability.</li>
                <li><strong>Rate limiting</strong> — With 3,443 datasets, bulk operations need pagination and caching.</li>
            </ul>
        </div>
    </div>

    <div class="footer">
        Generated by Serbian Data MCP investigation · data.gov.rs · {(summarize := __import__('datetime').datetime.now().strftime('%B %d, %Y'))}
    </div>
</div>
</body>
</html>"""

    index_path = EXPORT_DIR / "index.html"
    index_path.write_text(html, encoding="utf-8")
    logger.info(f"✅ Index page: {index_path}")


async def main() -> None:
    """Run all categories with incremental progress saving."""
    from serbian_data_mcp.api.client import UDataClient

    client = UDataClient()

    categories = [
        ("census", category_census),
        ("employment", category_employment),
        ("air_quality", category_air_quality),
        ("budgets", category_budgets),
        ("real_estate", category_real_estate),
        ("registries", category_registries),
        ("cross_analysis", category_cross_analysis),
        ("data_availability", category_municipal_data_availability),
    ]

    for name, fn in categories:
        try:
            info = await fn(client)
            save_progress(name, info)
            logger.info(f"  ✅ {name} complete")
        except Exception as e:
            logger.error(f"  ❌ {name} failed: {e}")
            save_progress(name, {"status": "failed", "error": str(e)})
        time.sleep(0.5)  # Rate limit

    await build_index_html()
    logger.info("=" * 60)
    logger.info("ALL DONE — Open exports/showcase/index.html")


if __name__ == "__main__":
    asyncio.run(main())
