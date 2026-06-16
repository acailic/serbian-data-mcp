"""Generate polished demo pages using MCP visualization tools with real Serbian data.

Creates standalone HTML files that demonstrate what the Serbian Data MCP
server can produce. Each demo uses real data fetched from data.gov.rs.

Demos:
  1. Employment Trends — line chart with forecasting
  2. Air Quality Dashboard — heatmap + bar ranking
  3. Census Changes — slope chart of municipality rankings
  4. Cross-Analysis — employment per household choropleth
  5. Census Infographic — full data story
  6. Index page linking to all demos
"""

from __future__ import annotations

import asyncio
import json
import logging
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import plotly.graph_objects as go

from serbian_data_mcp.api.client import UDataClient
from serbian_data_mcp.viz.charts import ChartBuilder
from serbian_data_mcp.viz.exporters import export_html
from serbian_data_mcp.viz.themes import apply_theme
from serbian_data_mcp.viz.novel_charts import slope_chart, waffle_chart
from serbian_data_mcp.viz.insights import extract_insights, generate_narrative
from serbian_data_mcp.viz.infographics import create_infographic
from serbian_data_mcp.viz.maps import SerbiaMapBuilder
from serbian_data_mcp.viz.forecast import forecast_linear
from serbian_data_mcp.viz.advanced_charts import AdvancedChartBuilder

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

DEMO_DIR = Path("exports/demo")
DEMO_DIR.mkdir(parents=True, exist_ok=True)


async def fetch_json_url(client: UDataClient, url: str) -> list[dict[str, Any]]:
    """Fetch JSON from URL with UTF-8 BOM handling."""
    ext = await client._get_external_client()
    resp = await ext.get(url)
    resp.raise_for_status()
    return json.loads(resp.content.decode("utf-8-sig"))


def write_demo_html(filename: str, title: str, body_html: str, extra_head: str = "") -> Path:
    """Write a standalone demo HTML page with dark styling."""
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title} — Serbian Data MCP</title>
    <script src="https://cdn.plot.ly/plotly-3.6.0.min.js" charset="utf-8"></script>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    {extra_head}
    <style>
        :root {{
            --bg: #0f1117; --bg-card: #1a1d28; --bg-card2: #222639;
            --text: #e8e8ed; --text-dim: #8b8d97; --text-muted: #555;
            --accent: #c62828; --accent2: #1565c0; --accent3: #ffab00; --accent4: #2a9d8f;
            --border: rgba(255,255,255,0.07); --radius: 12px;
        }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            background: var(--bg); color: var(--text);
            font-family: 'Inter', 'Segoe UI', system-ui, sans-serif;
            line-height: 1.65; min-height: 100vh;
            background-image: radial-gradient(ellipse at 15% 0%, rgba(21,101,192,0.06) 0%, transparent 50%),
                              radial-gradient(ellipse at 85% 100%, rgba(198,40,40,0.04) 0%, transparent 50%);
        }}
        .wrap {{ max-width: 1100px; margin: 0 auto; padding: 48px 24px; }}
        .header {{
            text-align: center; margin-bottom: 48px;
            padding-bottom: 32px; border-bottom: 1px solid var(--border);
        }}
        .header h1 {{
            font-size: 2.2rem; font-weight: 800; margin-bottom: 8px;
            background: linear-gradient(135deg, var(--text), var(--accent3));
            -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        }}
        .header .sub {{ color: var(--text-dim); font-size: 1rem; }}
        .badge {{
            display: inline-block; padding: 2px 10px; border-radius: 20px;
            font-size: 0.7rem; font-weight: 700; text-transform: uppercase;
            letter-spacing: 0.5px; background: var(--accent3); color: #0f1117;
            margin-bottom: 12px;
        }}
        .badge-blue {{ background: var(--accent2); color: white; }}
        .card {{
            background: var(--bg-card); border: 1px solid var(--border);
            border-radius: var(--radius); padding: 28px; margin-bottom: 24px;
        }}
        .card h2 {{
            font-size: 1.2rem; font-weight: 700; margin-bottom: 16px;
            color: var(--text);
        }}
        .stats {{
            display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
            gap: 16px; margin: 20px 0;
        }}
        .stat {{ text-align: center; padding: 16px; background: rgba(255,255,255,0.02); border-radius: 8px; }}
        .stat .num {{ font-size: 2rem; font-weight: 800; }}
        .stat .num.red {{ color: var(--accent); }}
        .stat .num.blue {{ color: var(--accent2); }}
        .stat .num.green {{ color: var(--accent4); }}
        .stat .num.gold {{ color: var(--accent3); }}
        .stat .lbl {{ color: var(--text-dim); font-size: 0.8rem; margin-top: 4px; }}
        .note {{ color: var(--text-muted); font-size: 0.78rem; margin-top: 8px; text-align: center; }}
        .nav {{
            text-align: center; margin-top: 40px; padding-top: 24px;
            border-top: 1px solid var(--border);
        }}
        .nav a {{
            color: var(--accent3); text-decoration: none; font-weight: 600;
        }}
        .nav a:hover {{ text-decoration: underline; }}
    </style>
</head>
<body>
<div class="wrap">
    <div class="header">
        <div class="badge">Serbian Data MCP</div>
        <h1>{title}</h1>
        <p class="sub">Real data from data.gov.rs · 3,443 datasets · 184 organizations</p>
    </div>
    {body_html}
    <div class="nav">
        <a href="index.html">← Back to Demos</a>
    </div>
</div>
</body>
</html>"""
    path = DEMO_DIR / filename
    path.write_text(html, encoding="utf-8")
    logger.info(f"  ✅ {path}")
    return path


def fig_to_embed_html(fig: go.Figure, div_id: str = "chart") -> str:
    """Convert Plotly figure to embedded HTML div."""
    raw = fig.to_html(full_html=False, include_plotlyjs=False, div_id=div_id)
    return raw


# =========================================================================
# Demo 1: Employment Trends
# =========================================================================


async def demo_employment(client: UDataClient) -> Path:
    """Employment trends with forecasting."""
    logger.info("Demo 1: Employment Trends")
    ds = await client.get_dataset("607fd7dd7de272771a0d3975")
    data = await fetch_json_url(client, ds.resources[0].url)

    # Extract yearly totals
    yearly = []
    for r in data:
        if r["IDTer"] == "RS" and r["IDModalitetRegZap"] == "0":
            yearly.append({"year": int(r["god"]), "total": r["vrednost"]})
    yearly.sort(key=lambda x: x["year"])

    # Employment types for latest year
    types = []
    latest_yr = max(r["god"] for r in data if r["IDTer"] == "RS" and r["IDModalitetRegZap"] == "0")
    for r in data:
        if r["IDTer"] == "RS" and r["god"] == latest_yr and r["IDModalitetRegZap"] != "0":
            types.append({"type": r["nModalitetRegZap"][:50], "value": r["vrednost"]})

    # Line chart
    builder = ChartBuilder(yearly)
    fig = builder.line_chart("year", "total", title="Registered Employment in Serbia (2015–2025)")
    fig = apply_theme(fig, "dark")
    chart1 = fig_to_embed_html(fig, "emp_trend")

    # Forecast
    forecast = forecast_linear(yearly, "year", "total", periods_ahead=5, method="linear")
    growth_rate = forecast.get("growth_rate", 0)
    r_sq = forecast.get("r_squared", 0)
    proj_note = forecast.get("projection_note", "")

    # Forecast line chart
    hist = forecast.get("historical_data", [])
    fore = forecast.get("forecast_data", [])
    trend = forecast.get("trend_line", [])

    all_years = hist + fore
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=[h["year"] for h in hist], y=[h["total"] for h in hist], mode="lines+markers", name="Actual", line=dict(color="#1565c0", width=3)))
    if fore:
        fig2.add_trace(go.Scatter(x=[f["year"] for f in fore], y=[f["_forecast"] for f in fore], mode="lines+markers", name="Forecast", line=dict(color="#ffab00", width=3, dash="dash")))
    if trend:
        fig2.add_trace(go.Scatter(x=[t["year"] for t in trend], y=[t["_trend"] for t in trend], mode="lines", name="Trend", line=dict(color="rgba(255,255,255,0.2)", width=1)))
    fig2.update_layout(title="Employment Forecast to 2030", **ChartBuilder(yearly).data.attrs.get("_layout_base", {}))
    fig2 = apply_theme(fig2, "dark")
    chart2 = fig_to_embed_html(fig2, "emp_forecast")

    # Pie chart of employment types
    builder_types = ChartBuilder(types)
    fig3 = builder_types.pie_chart("value", "type", title=f"Employment by Type ({latest_yr})")
    fig3 = apply_theme(fig3, "dark")
    chart3 = fig_to_embed_html(fig3, "emp_types")

    # Stats
    first = yearly[0]
    last = yearly[-1]
    pct_change = ((last["total"] / first["total"]) - 1) * 100

    body = f"""
    <div class="stats">
        <div class="stat"><div class="num green">{last['total']:,.0f}</div><div class="lbl">Employed ({latest_yr})</div></div>
        <div class="stat"><div class="num gold">+{pct_change:.1f}%</div><div class="lbl">Growth Since {first['year']}</div></div>
        <div class="stat"><div class="num blue">{r_sq:.3f}</div><div class="lbl">R² (Trend Fit)</div></div>
        <div class="stat"><div class="num">{growth_rate/1000:.1f}K</div><div class="lbl">Annual Growth</div></div>
    </div>
    <div class="card"><h2>📈 Employment Trend</h2>{chart1}</div>
    <div class="card"><h2>🔮 Forecast to 2030</h2>
        <p style="color:var(--text-dim);margin-bottom:12px;font-size:0.9rem">{proj_note}</p>
        {chart2}
    </div>
    <div class="card"><h2>📊 Employment by Type ({latest_yr})</h2>{chart3}</div>
    """
    return write_demo_html("01_employment.html", "Employment Trends", body)


# =========================================================================
# Demo 2: Air Quality Dashboard
# =========================================================================


async def demo_air_quality(client: UDataClient) -> Path:
    """Air quality dashboard with heatmap and ranking."""
    logger.info("Demo 2: Air Quality Dashboard")
    import openpyxl
    from io import BytesIO

    ds = await client.get_dataset("661909571df0e888307e3fa3")
    ext = await client._get_external_client()
    resp = await ext.get(ds.resources[0].url)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(BytesIO(resp.content), read_only=True)
    ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(h).strip() if h else "" for h in raw_rows[0]]
    stations = [h for h in headers[1:] if h]

    # Build tidy data for 2024
    tidy = []
    for row in raw_rows[1:]:
        if row[0] is None:
            continue
        date_str = str(row[0])
        for i, st in enumerate(stations):
            val = row[i + 1] if i + 1 < len(row) else None
            if val is not None and val != "" and val != 0:
                try:
                    tidy.append({"date": date_str, "station": st, "pm10": float(val)})
                except (ValueError, TypeError):
                    pass

    df = pd.DataFrame(tidy)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"])
    df["month"] = df["date"].dt.month

    # Monthly averages per station (heatmap)
    pivot = df.groupby(["station", "month"])["pm10"].mean().reset_index()
    top_stations = pivot.groupby("station")["pm10"].mean().nlargest(20).index.tolist()
    pivot_top = pivot[pivot["station"].isin(top_stations)]

    builder = AdvancedChartBuilder(pivot_top.to_dict("records"))
    fig1 = builder.heatmap("station", "month", "pm10", title="Monthly PM₁₀ by Station (2024) — Top 20 Polluted", theme="dark")
    chart1 = fig_to_embed_html(fig1, "heatmap")

    # Bar chart ranking — annual averages
    station_avg = df.groupby("station")["pm10"].mean().reset_index()
    station_avg.columns = ["station", "avg_pm10"]
    station_avg = station_avg.sort_values("avg_pm10", ascending=True)

    builder2 = ChartBuilder(station_avg.to_dict("records"))
    fig2 = builder2.bar_chart("station", "avg_pm10", title="Annual PM₁₀ Ranking (2024)", orientation="h")
    fig2.update_layout(height=max(600, len(station_avg) * 22))
    fig2 = apply_theme(fig2, "dark")
    chart2 = fig_to_embed_html(fig2, "ranking")

    # Monthly trend (average across all stations)
    monthly = df.groupby("month")["pm10"].mean().reset_index()
    monthly.columns = ["month", "avg_pm10"]
    month_names = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
                   7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}
    monthly["month_name"] = monthly["month"].map(month_names)

    builder3 = ChartBuilder(monthly.to_dict("records"))
    fig3 = builder3.bar_chart("month_name", "avg_pm10", title="Seasonal PM₁₀ Pattern (All Stations Average)")
    fig3 = apply_theme(fig3, "dark")
    chart3 = fig_to_embed_html(fig3, "seasonal")

    # Stats
    worst_st = station_avg.iloc[-1]
    best_st = station_avg.iloc[0]

    body = f"""
    <div class="stats">
        <div class="stat"><div class="num">{len(stations)}</div><div class="lbl">Monitoring Stations</div></div>
        <div class="stat"><div class="num">{len(df):,}</div><div class="lbl">Daily Measurements</div></div>
        <div class="stat"><div class="num red">{worst_st['avg_pm10']:.1f}</div><div class="lbl">Worst (µg/m³)<br><small>{worst_st['station']}</small></div></div>
        <div class="stat"><div class="num green">{best_st['avg_pm10']:.1f}</div><div class="lbl">Best (µg/m³)<br><small>{best_st['station']}</small></div></div>
    </div>
    <div class="card"><h2>🌡️ Monthly PM₁₀ Heatmap — Top 20 Stations</h2>{chart1}</div>
    <div class="card"><h2>📊 Station Ranking</h2>{chart2}</div>
    <div class="card"><h2>Winter vs Summer Pattern</h2>{chart3}
        <p style="color:var(--text-dim);margin-top:12px;font-size:0.85rem">
            Winter heating season (Jan–Feb) shows elevated PM₁₀, with cleaner air in summer months.
        </p>
    </div>
    """
    return write_demo_html("02_air_quality.html", "Air Quality Dashboard", body)


# =========================================================================
# Demo 3: Census Changes
# =========================================================================


async def demo_census(client: UDataClient) -> Path:
    """Census household changes with slope chart and waffle."""
    logger.info("Demo 3: Census Changes")
    ds22 = await client.get_dataset("65fc2fc669c6e53ef2b38c50")
    data22 = await fetch_json_url(client, ds22.resources[0].url)
    ds11 = await client.get_dataset("607fd62a7de272771a0d3842")
    data11 = await fetch_json_url(client, ds11.resources[0].url)

    def extract_totals(data):
        out = {}
        for r in data:
            if r["IDBrClDom"] == 0 and r["IDTipNaselja"] == "0" and r["IDTer"] != "RS":
                out[r["nTer"]] = r["vrednost"]
        return out

    munis22 = extract_totals(data22)
    munis11 = extract_totals(data11)

    # Build slope data (top/bottom by change)
    slope_rows = []
    for name in sorted(set(munis22.keys()) & set(munis11.keys())):
        slope_rows.append({"municipality": name, "hh_2011": munis11[name], "hh_2022": munis22[name]})

    df_slope = pd.DataFrame(slope_rows)
    df_slope["change"] = df_slope["hh_2022"] - df_slope["hh_2011"]
    df_slope = df_slope.sort_values("change")

    # Top losers and gainers for slope chart
    top = pd.concat([df_slope.head(8), df_slope.tail(8)])

    fig1 = slope_chart(top.to_dict("records"), "municipality", "hh_2011", "hh_2022",
                        title="Household Change 2011 → 2022 (Top Gainers & Losers)", theme="dark")
    fig1.update_layout(height=700)
    chart1 = fig_to_embed_html(fig1, "slope")

    # Household structure waffle (Serbia 2022)
    structure = []
    for r in data22:
        if r["nTer"] == "РЕПУБЛИКА СРБИЈА" and r["IDTipNaselja"] == "0" and r["IDBrClDom"] != 0:
            structure.append({"type": r["nBrClDom"], "count": r["vrednost"]})

    fig2 = waffle_chart(structure, "type", "count", title="Serbia Household Structure (2022 Census)", theme="dark")
    chart2 = fig_to_embed_html(fig2, "waffle")

    # Bar chart of top 10 gainers
    gainers = df_slope.tail(10).iloc[::-1]
    builder = ChartBuilder(gainers.to_dict("records"))
    fig3 = builder.bar_chart("municipality", "hh_2022", title="Top 10 Growing Municipalities (Households)", orientation="h")
    fig3 = apply_theme(fig3, "dark")
    chart3 = fig_to_embed_html(fig3, "gainers")

    serbia22 = next((r["vrednost"] for r in data22 if r["nTer"] == "РЕПУБЛИКА СРБИЈА" and r["IDBrClDom"] == 0 and r["IDTipNaselja"] == "0"), 0)
    serbia11 = next((r["vrednost"] for r in data11 if r["nTer"] == "РЕПУБЛИКА СРБИЈА" and r["IDBrClDom"] == 0 and r["IDTipNaselja"] == "0"), 0)
    change_pct = ((serbia22 / serbia11) - 1) * 100 if serbia11 else 0

    body = f"""
    <div class="stats">
        <div class="stat"><div class="num">{serbia22:,.0f}</div><div class="lbl">Households (2022)</div></div>
        <div class="stat"><div class="num blue">{serbia11:,.0f}</div><div class="lbl">Households (2011)</div></div>
        <div class="stat"><div class="num gold">+{change_pct:.1f}%</div><div class="lbl">Overall Change</div></div>
        <div class="stat"><div class="num">{len(slope_rows)}</div><div class="lbl">Municipalities</div></div>
    </div>
    <div class="card"><h2>📈 Slope Chart — Ranking Changes 2011→2022</h2>{chart1}</div>
    <div class="card"><h2>🏘️ Household Structure — Serbia 2022</h2>{chart2}</div>
    <div class="card"><h2>📊 Top Growing Municipalities</h2>{chart3}</div>
    """
    return write_demo_html("03_census.html", "Census Household Changes", body)


# =========================================================================
# Demo 4: Cross-Analysis (Employment per Household)
# =========================================================================


async def demo_cross_analysis(client: UDataClient) -> Path:
    """Cross-analysis: employment per household with map."""
    logger.info("Demo 4: Cross-Analysis")

    # Employment by municipality
    ds_emp = await client.get_dataset("607fd7da7de272771a0d3973")
    data_emp = await fetch_json_url(client, ds_emp.resources[0].url)

    emp_years = sorted(set(r["god"] for r in data_emp if r["IDTer"] != "RS"), reverse=True)
    latest = emp_years[0]

    emp_by_muni = {}
    for r in data_emp:
        if r["god"] == latest and r["IDTer"] != "RS":
            key = r["IDModalitetRegZap"] if "IDModalitetRegZap" in r else "0"
            if key == "0":
                if r["nTer"] not in emp_by_muni or r["vrednost"] > emp_by_muni[r["nTer"]]:
                    emp_by_muni[r["nTer"]] = r["vrednost"]

    # Census households
    ds_census = await client.get_dataset("65fc2fc669c6e53ef2b38c50")
    data_census = await fetch_json_url(client, ds_census.resources[0].url)

    hh_by_muni = {}
    for r in data_census:
        if r["IDBrClDom"] == 0 and r["IDTipNaselja"] == "0" and r["IDTer"] != "RS":
            hh_by_muni[r["nTer"]] = r["vrednost"]

    # Match
    cross = []
    for m in sorted(set(emp_by_muni.keys()) & set(hh_by_muni.keys())):
        ratio = emp_by_muni[m] / hh_by_muni[m] if hh_by_muni[m] else 0
        cross.append({"municipality": m, "employment": emp_by_muni[m], "households": hh_by_muni[m], "ratio": round(ratio, 2)})

    df = pd.DataFrame(cross).sort_values("ratio", ascending=False)

    # Top 20 bar chart
    top20 = df.head(20)
    builder = ChartBuilder(top20.to_dict("records"))
    fig1 = builder.bar_chart("municipality", "ratio", title=f"Employment per Household — Top 20 Municipalities ({latest})")
    fig1 = apply_theme(fig1, "dark")
    chart1 = fig_to_embed_html(fig1, "cross_bar")

    # Scatter: employment vs households
    builder2 = ChartBuilder(df.to_dict("records"))
    fig2 = builder2.scatter_plot("households", "employment", title="Employment vs Households (by municipality)")
    fig2 = apply_theme(fig2, "dark")
    chart2 = fig_to_embed_html(fig2, "cross_scatter")

    # Insights
    insights = extract_insights(df.to_dict("records"), entity_column="municipality")

    body = f"""
    <div class="stats">
        <div class="stat"><div class="num blue">{len(df)}</div><div class="lbl">Municipalities Matched</div></div>
        <div class="stat"><div class="num gold">{df['ratio'].mean():.2f}</div><div class="lbl">Avg Employment/Household</div></div>
        <div class="stat"><div class="num green">{df.iloc[0]['ratio']}</div><div class="lbl">Highest<br><small>{df.iloc[0]['municipality']}</small></div></div>
        <div class="stat"><div class="num red">{df.iloc[-1]['ratio']}</div><div class="lbl">Lowest<br><small>{df.iloc[-1]['municipality']}</small></div></div>
    </div>
    <div class="card"><h2>📊 Employment per Household — Top 20</h2>{chart1}</div>
    <div class="card"><h2>🔍 Employment vs Households (Scatter)</h2>{chart2}
        <p style="color:var(--text-dim);margin-top:12px;font-size:0.85rem">
            Points above the trend line have disproportionately high employment relative to household count.
        </p>
    </div>
    """
    return write_demo_html("04_cross_analysis.html", "Cross-Analysis: Employment × Demographics", body)


# =========================================================================
# Demo 5: Municipal Budgets
# =========================================================================


async def demo_budgets(client: UDataClient) -> Path:
    """Budget data availability overview."""
    logger.info("Demo 5: Municipal Budgets")

    # Scan budget datasets
    all_budgets = []
    page = 1
    while True:
        result = await client.search_datasets(query="Буџет", page_size=100, page=page)
        if not result.datasets:
            break
        for ds in result.datasets:
            org = ds.organization.name if ds.organization else "Unknown"
            tags = ds.tags or []
            import re
            years = re.findall(r"20\d{2}", ds.title)
            all_budgets.append({"title": ds.title, "org": org, "years": years, "tags": tags})
        if not result.has_next:
            break
        page += 1

    # Build treemap by organization
    from collections import Counter
    org_counts = Counter(b["org"] for b in all_budgets)
    treemap_data = [{"org": org, "count": count} for org, count in org_counts.most_common(15)]

    builder = AdvancedChartBuilder(treemap_data)
    fig1 = builder.treemap("org", "count", title="Budget Datasets by Municipality", theme="dark")
    chart1 = fig_to_embed_html(fig1, "budget_treemap")

    # Tags breakdown
    tag_counts = Counter(t for b in all_budgets for t in b["tags"])
    tag_data = [{"tag": t, "count": c} for t, c in tag_counts.most_common(10)]
    builder2 = ChartBuilder(tag_data)
    fig2 = builder2.bar_chart("tag", "count", title="Budget Dataset Tags", orientation="h")
    fig2 = apply_theme(fig2, "dark")
    chart2 = fig_to_embed_html(fig2, "tags")

    # Sample budget CSV data
    sample_csv = ""
    for b in all_budgets[:30]:
        if "csv" in str(b.get("formats", "")):
            try:
                ds = await client.get_dataset(b.get("id", ""))
                if ds and ds.resources:
                    res = ds.resources[0]
                    ext2 = await client._get_external_client()
                    resp = await ext2.get(res.url)
                    resp.raise_for_status()
                    lines = resp.text.split("\n")
                    sample_csv = "<br>".join(f'<span style="color:var(--text-muted);font-size:0.8rem">{l[:120]}</span>' for l in lines[:8])
                    break
            except Exception:
                continue

    body = f"""
    <div class="stats">
        <div class="stat"><div class="num">{len(all_budgets)}</div><div class="lbl">Budget Datasets</div></div>
        <div class="stat"><div class="num blue">{len(org_counts)}</div><div class="lbl">Municipalities</div></div>
        <div class="stat"><div class="num gold">XLSX</div><div class="lbl">Primary Format</div></div>
    </div>
    <div class="card"><h2>📊 Budget Datasets by Municipality</h2>{chart1}</div>
    <div class="card"><h2>🏷️ Tags & Categories</h2>{chart2}</div>
    <div class="card"><h2>📋 Sample Budget Data</h2>
        <div style="background:var(--bg);padding:16px;border-radius:8px;font-family:monospace;line-height:1.8;word-break:break-all">
            {sample_csv or '<span style="color:var(--text-muted)">Sample not available (XLSX format)</span>'}
        </div>
    </div>
    """
    return write_demo_html("05_budgets.html", "Municipal Budgets Overview", body)


# =========================================================================
# Demo 6: Real Estate Prices
# =========================================================================


async def demo_real_estate(client: UDataClient) -> Path:
    """Real estate prices overview."""
    logger.info("Demo 6: Real Estate")

    result = await client.search_datasets("непокретности", page_size=50)
    datasets = []
    for ds in result.datasets:
        org = ds.organization.name if ds.organization else "Unknown"
        import re
        years = re.findall(r"20\d{2}", ds.title)
        datasets.append({"title": ds.title, "org": org, "years": years})

    # Try to download national price data
    price_data = []
    for ds in result.datasets:
        if "ПРОСЕЧНИХ" in ds.title or "Просечн" in ds.title:
            try:
                d = await client.get_dataset(ds.id)
                if d.resources:
                    ext = await client._get_external_client()
                    resp = await ext.get(d.resources[0].url)
                    resp.raise_for_status()
                    import openpyxl
                    from io import BytesIO
                    wb = openpyxl.load_workbook(BytesIO(resp.content), read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    wb.close()
                    if len(rows) > 2:
                        headers = [str(h).strip() if h else "" for h in rows[0]]
                        for r in rows[1:]:
                            vals = [str(v)[:30] if v else "" for v in r]
                            price_data.append(dict(zip(headers, vals)))
                    break
            except Exception:
                continue

    # Table of datasets
    table_rows = ""
    for d in datasets[:12]:
        years_str = ", ".join(d["years"]) if d["years"] else "—"
        table_rows += f'<tr><td>{d["title"][:60]}</td><td>{d["org"][:30]}</td><td>{years_str}</td></tr>'

    # Sample data table
    sample_table = ""
    if price_data:
        headers_sample = list(price_data[0].keys())[:6]
        sample_table = "<tr>" + "".join(f"<th>{h}</th>" for h in headers_sample) + "</tr>"
        for row in price_data[:8]:
            sample_table += "<tr>" + "".join(f"<td>{row.get(h, '')}</td>" for h in headers_sample) + "</tr>"

    body = f"""
    <div class="stats">
        <div class="stat"><div class="num">{len(datasets)}</div><div class="lbl">Property Price Datasets</div></div>
        <div class="stat"><div class="num gold">{len(price_data)}</div><div class="lbl">Price Records (National)</div></div>
    </div>
    <div class="card"><h2>📋 Available Datasets</h2>
        <table style="width:100%;border-collapse:collapse;font-size:0.85rem;margin-top:12px">
            <tr><th style="color:var(--text-dim);text-align:left;padding:8px">Dataset</th><th style="color:var(--text-dim);text-align:left;padding:8px">Publisher</th><th style="color:var(--text-dim);text-align:left;padding:8px">Years</th></tr>
            {table_rows}
        </table>
    </div>
    {"<div class='card'><h2>💱 National Price Data Sample</h2><table style='width:100%;border-collapse:collapse;font-size:0.85rem;margin-top:12px'>" + sample_table + "</table></div>" if sample_table else ""}
    """
    return write_demo_html("06_real_estate.html", "Real Estate Prices", body)


# =========================================================================
# Demo Index Page
# =========================================================================


async def build_index() -> Path:
    """Build index page linking to all demos."""
    demos = [
        {"file": "01_employment.html", "title": "Employment Trends", "desc": "10-year registered employment data with forecasting to 2030", "badge": "HIGH", "badge_class": ""},
        {"file": "02_air_quality.html", "title": "Air Quality Dashboard", "desc": "Daily PM₁₀ from 39 stations — heatmap, ranking, seasonal patterns", "badge": "HIGH", "badge_class": ""},
        {"file": "03_census.html", "title": "Census Changes", "desc": "Household changes 2011→2022 with slope chart and waffle diagram", "badge": "HIGH", "badge_class": ""},
        {"file": "04_cross_analysis.html", "title": "Cross-Analysis", "desc": "Employment per household across 200+ municipalities", "badge": "MEDIUM", "badge_class": "badge-blue"},
        {"file": "05_budgets.html", "title": "Municipal Budgets", "desc": "Budget data availability from 19 municipalities", "badge": "HIGH", "badge_class": ""},
        {"file": "06_real_estate.html", "title": "Real Estate Prices", "desc": "Property price datasets and tax assessment data", "badge": "HIGH", "badge_class": ""},
    ]

    cards = ""
    for d in demos:
        cards += f"""
        <a href="{d['file']}" style="text-decoration:none;color:inherit;display:block">
            <div class="card" style="transition:transform 0.15s,border-color 0.15s;cursor:pointer" onmouseover="this.style.borderColor='var(--accent3)';this.style.transform='translateY(-2px)'" onmouseout="this.style.borderColor='var(--border)';this.style.transform='none'">
                <div class="{'badge badge-blue' if d['badge_class'] else 'badge'}" style="margin-bottom:12px">{d['badge']}</div>
                <h2 style="margin-bottom:8px">{d['title']}</h2>
                <p style="color:var(--text-dim);font-size:0.9rem">{d['desc']}</p>
                <p style="color:var(--accent3);font-size:0.8rem;margin-top:12px;font-weight:600">Open demo →</p>
            </div>
        </a>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Serbian Data MCP — Demo Pages</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg: #0f1117; --bg-card: #1a1d28;
            --text: #e8e8ed; --text-dim: #8b8d97;
            --accent: #c62828; --accent2: #1565c0; --accent3: #ffab00; --accent4: #2a9d8f;
            --border: rgba(255,255,255,0.07); --radius: 12px;
        }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            background: var(--bg); color: var(--text);
            font-family: 'Inter', 'Segoe UI', system-ui, sans-serif;
            line-height: 1.65; min-height: 100vh;
            background-image: radial-gradient(ellipse at 15% 0%, rgba(21,101,192,0.06) 0%, transparent 50%),
                              radial-gradient(ellipse at 85% 100%, rgba(198,40,40,0.04) 0%, transparent 50%);
        }}
        .wrap {{ max-width: 800px; margin: 0 auto; padding: 48px 24px; }}
        .header {{ text-align: center; margin-bottom: 48px; padding-bottom: 32px; border-bottom: 1px solid var(--border); }}
        .header h1 {{
            font-size: 2.4rem; font-weight: 800; margin-bottom: 8px;
            background: linear-gradient(135deg, #c62828, #ffab00, #1565c0);
            -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        }}
        .header .sub {{ color: var(--text-dim); font-size: 1rem; margin-top: 8px; }}
        .header .mcp {{
            display: inline-block; margin-top: 16px; padding: 8px 20px;
            background: var(--bg-card); border: 1px solid var(--border);
            border-radius: 8px; font-family: monospace; font-size: 0.85rem; color: var(--accent3);
        }}
        .badge {{
            display: inline-block; padding: 2px 10px; border-radius: 20px;
            font-size: 0.7rem; font-weight: 700; text-transform: uppercase;
            letter-spacing: 0.5px; background: var(--accent3); color: #0f1117;
        }}
        .badge-blue {{ background: var(--accent2); color: white; }}
        .card {{
            background: var(--bg-card); border: 1px solid var(--border);
            border-radius: var(--radius); padding: 28px; margin-bottom: 20px;
        }}
        .card h2 {{ font-size: 1.2rem; font-weight: 700; margin-bottom: 0; }}
        .footer {{ text-align: center; color: var(--text-dim); margin-top: 48px; font-size: 0.8rem; padding-top: 24px; border-top: 1px solid var(--border); }}
    </style>
</head>
<body>
<div class="wrap">
    <div class="header">
        <h1>🇷🇸 Serbian Data MCP — Demos</h1>
        <p class="sub">Real visualizations from real data on data.gov.rs</p>
        <p class="sub">3,443 datasets · 184 organizations · Census · Employment · Air Quality · Budgets</p>
        <div class="mcp">pip install serbian-data-mcp</div>
    </div>
    {cards}
    <div class="footer">
        Generated with Serbian Data MCP · data.gov.rs · All data is public open data from the Republic of Serbia
    </div>
</div>
</body>
</html>"""

    path = DEMO_DIR / "index.html"
    path.write_text(html, encoding="utf-8")
    logger.info(f"  ✅ {path}")
    return path


# =========================================================================
# Main
# =========================================================================


async def main() -> None:
    """Generate all demo pages."""
    client = UDataClient()

    demos = [
        demo_employment,
        demo_air_quality,
        demo_census,
        demo_cross_analysis,
        demo_budgets,
        demo_real_estate,
    ]

    for fn in demos:
        try:
            await fn(client)
        except Exception as e:
            logger.error(f"  ❌ {fn.__name__} failed: {e}")

    await build_index()

    # Also copy to main exports
    import shutil
    for f in DEMO_DIR.iterdir():
        if f.suffix == ".html":
            shutil.copy2(f, Path("exports") / f.name)

    logger.info("=" * 60)
    logger.info("ALL DEMOS DONE — Open exports/demo/index.html")


if __name__ == "__main__":
    asyncio.run(main())
